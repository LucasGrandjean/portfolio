
import csv
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, Reference
from openpyxl.drawing.image import Image
from openpyxl.chart.axis import DateAxis
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import PatternFill
from openpyxl.chart.label import DataLabel
from openpyxl.worksheet.table import Table, TableColumn, range_boundaries, TableStyleInfo
import codecs
import os
import shutil
from io import BytesIO

def filter_csv_by_sites(csv_file, selected_sites):
    csv_data = csv_file.read().decode('utf-8')
    csv_lines = csv_data.split('\n')
    csv_reader = csv.DictReader(csv_lines, delimiter=';')

    filtered_data = [row for row in csv_reader if row['Site'] in selected_sites]
    return filtered_data

def creerBilanSynthese(data):
    merged_data = []
    for site_data in data.values():
        merged_data.extend(site_data)
    return merged_data

def formatLieu(site):
    return site.replace(" ", "_").replace("/", "_").replace("\\", "_")
    
def genererNomFichier(prefix, site):
    return f"{prefix}_{formatLieu(site)}.xlsx"
    
def faireDonneesSexe(fileCSV):
    sommeHomme = 0
    sommeFemme = 0
    for row in fileCSV:
        genre = row['Genre']
        if genre == 'Homme':
            sommeHomme += 1
        elif genre == 'Femme':
            sommeFemme += 1
    return sommeHomme, sommeFemme

def faireDonneesDemarches(fileCSV):
    donneesOperateur = defaultdict(lambda: {'occurrences': 0, 'typeTotal': 0})
    for row in fileCSV:
        actionOperateur = row['OperateurAction']
        type_reception = row['Type Reception']
        operateur = row['Operateur']
        types = type_reception.split(', ')
        occurrences = donneesOperateur[actionOperateur]['occurrences']
        typeTotal = donneesOperateur[actionOperateur]['typeTotal']
        donneesOperateur[actionOperateur]['occurrences'] = occurrences + 1
        donneesOperateur[actionOperateur]['typeTotal'] = typeTotal + len(types)
        donneesOperateur[actionOperateur]['operateur'] = operateur
    return donneesOperateur

def faireDonneesOperateurs(fileCSV):
    AggregatOperateurs = set()
    SommeAggregatOperateurs = defaultdict(lambda: {'count': 0, 'typeTotal': 0})
    for row in fileCSV:
        operateur = row['Operateur']
        types = row['Type Reception'].split(', ')
        AggregatOperateurs.add(operateur)
        SommeAggregatOperateurs[operateur]['count'] += 1
        SommeAggregatOperateurs[operateur]['typeTotal'] += len(types)
    return AggregatOperateurs, SommeAggregatOperateurs

def faireDonneesDuree(fileCSV):
    sommeDuree = defaultdict(int)
    for row in fileCSV:
        duree = row['Durée']
        sommeDuree[duree] += 1
    return sommeDuree

def nettoyerNomFichier(filename):
    replacements = ['Pimms_Melun_-_', 'PM_Melun_-_', 'Permanence_', 'PERMANENCE_', 'permanence_']
    for replacement in replacements:
        filename = filename.replace(replacement, '')
    return filename

def doExtract(fileCSV, output_filename, isSummary=0):

    sommeHomme, sommeFemme = faireDonneesSexe(fileCSV)
    donneesOperateur = faireDonneesDemarches(fileCSV)
    AggregatOperateurs, SommeAggregatOperateurs = faireDonneesOperateurs(fileCSV)
    sommeDuree = faireDonneesDuree(fileCSV)

    # Création du Excel
    workbook = Workbook()

    # Création d'un Excel pour Homme et Femme
    tableauSexe = workbook.create_sheet(title='Sexe')
    tableauSexe.append(['Sexe', 'Total'])
    tableauSexe.append(['Homme', sommeHomme])
    tableauSexe.append(['Femme', sommeFemme])

    # Création d'un Excel pour les démarches par opérateur
    tableauDemarcheOperateur = workbook.create_sheet(title='Par démarche')
    tableauDemarcheOperateur.append(['Démarche', 'Total d’usagers', 'Motifs'])
    for actionOperateur, data in donneesOperateur.items():
        if actionOperateur == "1ère demande de logement social sur www.demande-logement-social.gouv.fr":
            actionOperateur = "Demande de logement social"
        if actionOperateur == "Création du compte www.demande-logement-social.gouv.fr":
            actionOperateur = "Création compte Logement Social"
        elif actionOperateur in ["Création de compte", "Consultation compte", "Accompagnement", "Explication Courriers", "Changement de situation/coordonnées", "Appel - Info - Conseil", "Création compte en ligne",  "Récupération ID et mot de passe", "Autre", "Prise de RDV", "Accompagnement à la complétude du dossier", "Accompagnement démarche", "Accompagnement démarches en ligne", "Accompagnement démarches", "Accompagnement démarche en ligne", "Accompagnement à la complétude de dossier", "Création de compte en ligne", "Informations - Orientation", "Information - Orientation", "Explication de courrier", "Changement de domiciliation"]:
            operateur = data.get('operateur', '') 
            if operateur == "Ministère des finances et comptes publics / DGFIP" :
                operateur = "DGFIP"
            if operateur == "Assurance retraite / CARSAT - CNAV":
                operateur = "Retraite"
            if operateur == "Ma demande de logement social":
                operateur = "Logement social"
            if operateur == "Maison départementale des personnes handicapées (MDPH)":
                operateur = "MDPH"
            if operateur == "Ministère de l'Intérieur - Préfecture":
                operateur = "Préfecture"
            actionOperateur = f"{actionOperateur} ({operateur})"
        occurrences = data['occurrences']
        typeTotal = data['typeTotal']
        operateur = data.get('operateur', '')  
        tableauDemarcheOperateur.append([actionOperateur, occurrences, typeTotal])

    # Création des agrégats d'opérateurs
    tableauAgregatOperateurs = workbook.create_sheet(title='Par Opérateur')
    tableauAgregatOperateurs.append(['Opérateur', 'Total d\'usagers', 'Motifs'])
    for operator, data in SommeAggregatOperateurs.items():
        count = data['count']
        typeTotal = data['typeTotal']
        tableauAgregatOperateurs.append([operator, count, typeTotal])

    # Création d'un tableau par durée
    tableauDuree = workbook.create_sheet(title='Par durée')
    tableauDuree.append(['Temps', 'Total'])
    for duree, count in sommeDuree.items():
        tableauDuree.append([duree, count])

    # Suppression du premier tableau vide créee par défaut
    default_sheet = workbook['Sheet']
    workbook.remove(default_sheet)

    # Sauvegarde du Excel
    output_filename = nettoyerNomFichier(output_filename)
    workbook.save(output_filename)
    return output_filename  
   
def summary_mediateur(fileCSV, output_filename):
    # Load CSV data and create output Excel workbook
    workbook = Workbook()
    csv_data = defaultdict(lambda: defaultdict(lambda: {'unique_identifiers': set(), 'occurrences': 0}))
    contact_csv = defaultdict(lambda: defaultdict(lambda: {'unique_identifiers': set(), 'occurrences': 0}))

    for row in fileCSV:
        nom = row['Nom']
        identifiant = row['Identifiant']
        site = row['Site']
        contact = row['Contact']

        csv_data[nom][site]['unique_identifiers'].add(identifiant)
        csv_data[nom][site]['occurrences'] += 1
            
        contact_csv[nom][contact]['unique_identifiers'].add(identifiant)
        contact_csv[nom][contact]['occurrences'] += 1

    # Create summary sheets
    summary_nom_sheet = workbook.create_sheet(title='Par Nom')
    summary_nom_sheet.append(['Nom', 'Usagers', 'Démarches'])

    summary_site_sheet = workbook.create_sheet(title='Par Permanence')
    summary_site_sheet.append(['Nom', 'Permanence', 'Usagers', 'Démarches'])

    summary_contact_sheet = workbook.create_sheet(title='Par Contact')
    summary_contact_sheet.append(['Nom', 'Type de Contact', 'Usagers', 'Démarches'])

    for nom, sites_data in csv_data.items():
        total_unique_identifiers = sum(len(data['unique_identifiers']) for data in sites_data.values())
        total_occurrences = sum(data['occurrences'] for data in sites_data.values())
        summary_nom_sheet.append([nom, total_unique_identifiers, total_occurrences])

        for site, data in sites_data.items():
            unique_identifiers_count = len(data['unique_identifiers'])
            occurrences = data['occurrences']
            summary_site_sheet.append([nom, site, unique_identifiers_count, occurrences])

    for nom, contacts_data in contact_csv.items():

        for contact, data in contacts_data.items():
            unique_identifiers_count = len(data['unique_identifiers'])
            occurrences = data['occurrences']
            summary_contact_sheet.append([nom, contact, unique_identifiers_count, occurrences])

    # Remove the default sheet created by openpyxl
    default_sheet = workbook['Sheet']
    workbook.remove(default_sheet)

    # Save the workbook
    workbook.save(output_filename)
    return output_filename